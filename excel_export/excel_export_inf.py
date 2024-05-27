import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
import pandas as pd
from parameters.parametres_inf import I, J, K  # , part_time_I, full_time_I, nb_part_time_agents


def to_excel(values, variable_names, dest_path="export/nurse_schedule.xlsx"):

    # Create the DataFrame structure
    data = []
    for i in I:  # For each nurse
        row = []
        for j in J:  # For each day
            for k in K:  # For each shift
                var_name = f"x{i},{j},{k}"
                index = variable_names.index(var_name)
                row.append(int(values[index]))  # Convert float to int for clarity
        data.append(row)

    # Convert list to DataFrame
    shifts = {1: 'M', 2: 'S', 3: 'T'}
    days = {1: 'Lundi', 2: 'Mardi', 3: 'Mercredi', 4: 'Jeudi', 5: 'Vendredi', 6: 'Week-end'}
    columns = [f"Semaine_{(j//3) // 6 + 1} {days[(j//3) % 6 + 1]} {shifts[k]}"
               for j in range(0, 3*len(J), 3) for k in K]
    df = pd.DataFrame(data, index=[f"Agent {i}" for i in I], columns=columns)

    # Duplicate the week-end columns into two separate columns for Samedi and Dimanche
    # First replace "Week-end" with "Samedi"
    df.columns = df.columns.str.replace('Week-end', 'Samedi')
    # Then duplicate the columns for Dimanche, inserting them after Samedi
    weeks = max(J) // 6
    for w in range(weeks):
        for i, c in enumerate(df.columns):
            if 'Samedi' in c and f"Semaine_{w+1} " in c:
                df.insert(i+3, c.replace('Samedi', 'Dimanche'), df[c])

    # Adding the total number of shifts per nurse
    df['Total Shifts'] = df.sum(axis=1)

    # Add a row for the number of nurses assigned per shift
    total_nurses = df.sum(axis=0)
    total_nurses.name = 'Total Agents'
    df = pd.concat([df, total_nurses.to_frame().T])
    # Save as excel
    df.to_excel(dest_path)


def openpyxl_formatting(src_path="export/nurse_schedule_v1.xlsx",
                        dest_path="export/nurse_schedule_openpyxl_v1.xlsx"):

    wb = load_workbook(src_path)
    ws = wb.active

    # Assuming your data starts from column B (column A might be the index)
    # And that the 'Week X' formatting needs to apply from column B onward
    # Insert new rows for Week and Day headers
    ws.insert_rows(1)
    ws.insert_rows(1)

    # Apply header information
    header_row = 1  # Week headers
    day_header_row = 2  # Day headers

    # Set the initial week and day to manage merging
    current_week = None
    current_day = None
    week_start_col = 2
    day_start_col = 2

    for col in range(2, ws.max_column):
        # Extract week and day from the original cell
        original_header = ws.cell(row=3, column=col).value  # assuming original headers start from row 3 now
        if original_header:
            week, day, shift = original_header.split(" ")
            ws.cell(row=header_row, column=col).value = week
            ws.cell(row=day_header_row, column=col).value = day
            ws.cell(row=3, column=col).value = shift  # Shift goes back to row 3

            # Manage merging for week
            if week != current_week:
                if current_week is not None:
                    ws.merge_cells(start_row=header_row, start_column=week_start_col,
                                   end_row=header_row, end_column=col-1)
                current_week = week
                week_start_col = col

            # Manage merging for day
            if day != current_day:
                if current_day is not None:
                    ws.merge_cells(start_row=day_header_row, start_column=day_start_col,
                                   end_row=day_header_row, end_column=col-1)
                current_day = day
                day_start_col = col

    # Final merge for the last week and day
    if current_week:
        ws.merge_cells(start_row=header_row, start_column=week_start_col,
                       end_row=header_row, end_column=ws.max_column-1)
    if current_day:
        ws.merge_cells(start_row=day_header_row, start_column=day_start_col,
                       end_row=day_header_row, end_column=ws.max_column-1)

    # Set column widths and general styles as before
    for i, col in enumerate(ws.columns):
        # Skip first and last columns
        if i == 0 or i == ws.max_column-1:
            continue
        ws.column_dimensions[get_column_letter(col[0].column)].width = 5

    # Apply bold and centered style to day names
    bold_centered = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    for col in range(2, ws.max_column + 1):
        for row in [header_row, day_header_row]:
            cell = ws.cell(row=row, column=col)
            cell.font = bold_centered
            cell.alignment = center_alignment

    # Add a row with day-wise totals at the end
    total_row = ws.max_row
    new_row = ws.max_row + 1
    ws.insert_rows(new_row)
    ws.cell(row=new_row, column=1).value = "Total day-wise"
    ws.cell(row=new_row, column=1).font = bold_centered
    ws.cell(row=new_row, column=1).alignment = center_alignment
    for col in range(2, ws.max_column, 3):
        ws.cell(row=new_row, column=col).value = f"=SUM({get_column_letter(col)}{total_row}:{get_column_letter(col+2)}{total_row})"  # noqa
        ws.cell(row=new_row, column=col).font = bold_centered
        ws.cell(row=new_row, column=col).alignment = center_alignment
    # Group the total values for each day together
    for col in range(2, ws.max_column, 3):
        ws.merge_cells(start_row=new_row, start_column=col, end_row=new_row, end_column=col+2)

    # Define different border styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    # Apply thin borders to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Save the changes
    wb.save(dest_path)


def to_excel_v2(values, variable_names, dest_path="export/nurse_schedule_v2.xlsx"):
    """The excel file here will have only one value per day, with a string value for the shift type."""
    shifts = {1: 'M', 2: 'S', 3: 'T'}
    # Create the DataFrame structure
    data = []
    for i in I:  # For each nurse
        row = []
        for j in J:  # For each day
            shift_value = "R"  # Rest by default
            for k in K:  # For each shift
                index = variable_names.index(f"x{i},{j},{k}")
                if values[index] == 1:
                    shift_value = shifts[k]
            row.append(shift_value)
        data.append(row)

    # Convert list to DataFrame
    days = {1: 'Lundi', 2: 'Mardi', 3: 'Mercredi', 4: 'Jeudi', 5: 'Vendredi', 6: 'Week-end'}
    columns = [f"Semaine_{(j//3) // 6 + 1} {days[(j//3) % 6 + 1]}"
               for j in range(0, 3*len(J), 3)]
    index_names = []
    for i in I:
        # if i in part_time_I:
        #     index_names.append(f"Agent {i} (80%)")
        # else:
        index_names.append(f"Agent {i} (100%)")
    df = pd.DataFrame(data, index=index_names, columns=columns)

    # Duplicate the week-end columns into two separate columns for Samedi and Dimanche
    # First replace "Week-end" with "Samedi"
    df.columns = df.columns.str.replace('Week-end', 'Samedi')
    # Then duplicate the columns for Dimanche, inserting them after Samedi
    weeks = max(J) // 6
    for w in range(weeks):
        for i, c in enumerate(df.columns):
            if 'Samedi' in c and f"Semaine_{w+1} " in c:
                df.insert(i+1, c.replace('Samedi', 'Dimanche'), df[c])

    # Add a row for the number of nurses assigned per shift (excluding rest days)
    df['Total'] = df.apply(lambda row: sum([1 for shift in row if shift in ["M", "S", "T"]]), axis=1)
    total_nurses = df.apply(lambda col: sum([1 for shift in col if shift in ["M", "S", "T"]]), axis=0)
    total_nurses.name = 'Total'
    df = pd.concat([df, total_nurses.to_frame().T])
    for shift in shifts.values():
        df[f'Total {shift}'] = df.apply(lambda row: sum([1 for s in row if s == shift]), axis=1)
        total_nurses = df.apply(lambda col: sum([1 for s in col if s == shift]), axis=0)
        total_nurses.name = f'Total {shift}'
        df = pd.concat([df, total_nurses.to_frame().T])
    df['Total R'] = df.apply(lambda row: sum([1 for shift in row if shift == "R"]), axis=1)
    # Save as excel
    df.to_excel(dest_path)


def openpyxl_formatting_v2(src_path="export/nurse_schedule_v2.xlsx",
                           dest_path="export/nurse_schedule_openpyxl_v2.xlsx"):
    wb = load_workbook(src_path)
    ws = wb.active

    # Insert new rows for Week and Day headers
    ws.insert_rows(1)
    # Apply header information
    header_row = 1  # Week headers
    # Set the initial week and day to manage merging
    current_week = None
    week_start_col = 2
    week_dict = {"Lundi": "L", "Mardi": "M", "Mercredi": "Me", "Jeudi": "J",
                 "Vendredi": "V", "Samedi": "S", "Dimanche": "D"}

    for col in range(2, ws.max_column-4):
        # Extract week and day from the original cell
        original_header = ws.cell(row=2, column=col).value  # assuming original headers start from row 3 now
        if original_header:
            week, day = original_header.split(" ")
            ws.cell(row=header_row, column=col).value = week
            ws.cell(row=header_row+1, column=col).value = week_dict[day]
            # Manage merging for week
            if week != current_week:
                if current_week is not None:
                    ws.merge_cells(start_row=header_row, start_column=week_start_col,
                                   end_row=header_row, end_column=col-1)
                current_week = week
                week_start_col = col

    # Final merge for the last week and day
    if current_week:
        ws.merge_cells(start_row=header_row, start_column=week_start_col,
                       end_row=header_row, end_column=ws.max_column-5)

    # Set column widths and general styles as before
    for i, col in enumerate(ws.columns):
        # Skip first and last columns
        if i == 0 or i > ws.max_column-6:
            continue
        ws.column_dimensions[get_column_letter(col[0].column)].width = 5
    # First col should be wider
    ws.column_dimensions['A'].width = 15

    # Apply bold and centered style to day names
    bold_centered = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    for col in range(2, ws.max_column + 1):
        for row in [header_row, header_row+1]:
            cell = ws.cell(row=row, column=col)
            cell.font = bold_centered
            cell.alignment = center_alignment

    # Define different border styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    # Apply thin borders to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Paint the header rows (week and day) with in grey and light grey
    for col in range(2, ws.max_column + 1):
        for row in [header_row, header_row+1]:
            cell = ws.cell(row=row, column=col)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Paint the four final rows with a light grey background
    for row in range(ws.max_row-3, ws.max_row+1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    # Total row values in bold
    for col in range(2, ws.max_column - 4):
        ws.cell(row=ws.max_row-3, column=col).font = bold_centered

    # Values with "R" in light blue
    for row in range(2, ws.max_row - 3):
        for col in range(2, ws.max_column - 4):
            cell = ws.cell(row=row, column=col)
            if cell.value == "R":
                cell.fill = openpyxl.styles.PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")

    # Last five columns in light grey
    for col in range(ws.max_column - 4, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    # Col Total in bold
    for row in range(1, ws.max_row-3):
        ws.cell(row=row, column=ws.max_column-4).font = bold_centered

    # # Paint in pink the first shift of first agent at 80%
    # for i in part_time_I:
    #     ws.cell(row=2+i, column=2+7*(i-1)).fill = openpyxl.styles.PatternFill(start_color="FFC0CB",
    #                                                                           end_color="FFC0CB",
    #                                                                           fill_type="solid")
    # # Paint in orange the first shift of first agent at 100%
    # for i in full_time_I:
    #     ws.cell(row=2+i, column=2+7*(i-nb_part_time_agents-1)).fill = openpyxl.styles.PatternFill(start_color="FFA500",  # noqa
    #                                                                                               end_color="FFA500",
    #                                                                                               fill_type="solid")

    # Paint in orange the first shift of first agent
    multiple_6 = [j for j in J if j % 6 == 0]
    not_multiple_6 = [j for j in J if j % 6 != 0]
    nb_value_cols = 2 * len(multiple_6) + len(not_multiple_6)

    for i in I:
        ws.cell(row=2+i, column=(2+(7*(i-1))) % nb_value_cols).fill = openpyxl.styles.PatternFill(
                                                                                                start_color="FFA500",
                                                                                                end_color="FFA500",
                                                                                                fill_type="solid"
                                                                                                )
    # Save the changes
    wb.save(dest_path)
