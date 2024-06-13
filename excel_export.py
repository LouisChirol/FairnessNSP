import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
import os
import pandas as pd


def to_excel(
        values, variable_names,
        I, J, K, part_time_I, nb_shifts,
        dest_path="output/nurses_schedule.xlsx"
             ):
    """The excel file here will have only one value per day, with a string value for the shift type."""
    shifts = {1: 'M', 2: 'S', 3: 'T'}
    # Create the DataFrame structure
    data = []
    for i in I:  # For each nurse
        row = []
        for j in J:  # For each day
            shift_value = "R"  # Rest by default
            for k in K[:nb_shifts]:  # For each shift
                index = variable_names.index(f"x{i},{j},{k}")
                if values[index] == 1:
                    shift_value = shifts[k]
                if values[variable_names.index(f"x{i},{j},{k + nb_shifts}")] == 1:
                    shift_value += " (RA)"
            row.append(shift_value)
        data.append(row)

    # Convert list to DataFrame
    days = {1: 'Lundi', 2: 'Mardi', 3: 'Mercredi', 4: 'Jeudi', 5: 'Vendredi', 6: 'Week-end'}
    columns = [f"Semaine_{(j // 3) // 6 + 1} {days[(j // 3) % 6 + 1]}"
               for j in range(0, 3 * len(J), 3)]
    index_names = []
    for i in I:
        if i in part_time_I:
            index_names.append(f"Agent {i} (80%)")
        else:
            index_names.append(f"Agent {i} (100%)")
    df = pd.DataFrame(data, index=index_names, columns=columns)

    # Duplicate the week-end columns into two separate columns for Samedi and Dimanche
    # First replace "Week-end" with "Samedi"
    df.columns = df.columns.str.replace('Week-end', 'Samedi')
    # Then duplicate the columns for Dimanche, inserting them after Samedi
    weeks = max(J) // 6
    for w in range(weeks):
        for i, c in enumerate(df.columns):
            if 'Samedi' in c and f"Semaine_{w + 1} " in c:
                df.insert(i+1, c.replace('Samedi', 'Dimanche'), df[c])

    # Add a row for the number of nurses assigned per shift (excluding rest days)
    shift_names = ["M", "S", "T"]
    df['Total'] = df.apply(lambda row: sum([1 for shift in row if shift in shift_names]), axis=1)
    total_nurses = df.apply(lambda col: sum([1 for shift in col if shift in shift_names]), axis=0)
    total_nurses.name = 'Total'
    df = pd.concat([df, total_nurses.to_frame().T])
    for shift in ["M", "S", "T"]:
        df[f'Total {shift}'] = df.apply(lambda row: sum([1 for s in row if s == shift]), axis=1)
        total_nurses = df.apply(lambda col: sum([1 for s in col if s == shift]), axis=0)
        total_nurses.name = f'Total {shift}'
        df = pd.concat([df, total_nurses.to_frame().T])
    df['Total R'] = df.apply(lambda row: sum([1 for shift in row if shift == "R"]), axis=1)

    # Save as excel
    df.to_excel(dest_path)


def openpyxl_formatting(
        I, J,
        src_path="output/nurses_schedule.xlsx",
        dest_path="output/nurses_schedule.xlsx"
        ):
    wb = load_workbook(src_path)
    ws = wb.active

    # Insert new rows for Week and Day headers
    ws.insert_rows(1)
    # Apply header information
    header_row = 1  # Week headers
    # Set the initial week and day to manage merging
    current_week = None
    week_start_col = 2
    week_dict = {"Lundi": "L", "Mardi": "M", "Mercredi": "Me", "Jeudi": "J",
                 "Vendredi": "V", "Samedi": "S", "Dimanche": "D"}

    for col in range(2, ws.max_column - 4):
        # Extract week and day from the original cell
        original_header = ws.cell(row=2, column=col).value  # assuming original headers start from row 3 now
        if original_header:
            week, day = original_header.split(" ")
            ws.cell(row=header_row, column=col).value = week
            ws.cell(row=header_row + 1, column=col).value = week_dict[day]
            # Manage merging for week
            if week != current_week:
                if current_week is not None:
                    ws.merge_cells(start_row=header_row, start_column=week_start_col,
                                   end_row=header_row, end_column=(col - 1))
                current_week = week
                week_start_col = col

    # Final merge for the last week and day
    if current_week:
        ws.merge_cells(start_row=header_row, start_column=week_start_col,
                       end_row=header_row, end_column=(ws.max_column - 5))

    # Set column widths and general styles as before
    for i, col in enumerate(ws.columns):
        # Skip first and last columns
        if i == 0 or i > (ws.max_column - 6):
            continue
        ws.column_dimensions[get_column_letter(col[0].column)].width = 6
    # First col should be wider
    ws.column_dimensions['A'].width = 15

    # Apply bold and centered style to day names
    bold_centered = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    for col in range(2, ws.max_column + 1):
        for row in [header_row, header_row + 1]:
            cell = ws.cell(row=row, column=col)
            cell.font = bold_centered
            cell.alignment = center_alignment

    # Define different border styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    # Apply thin borders to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Paint the header rows (week and day) with in grey and light grey
    for col in range(2, ws.max_column + 1):
        for row in [header_row, header_row + 1]:
            cell = ws.cell(row=row, column=col)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Paint the four final rows with a light grey background
    for row in range(ws.max_row - 3, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    # Total row values in bold
    for col in range(2, ws.max_column - 4):
        ws.cell(row=ws.max_row - 3, column=col).font = bold_centered

    # Values with "R" in light blue
    for row in range(2, ws.max_row - 3):
        for col in range(2, ws.max_column - 4):
            cell = ws.cell(row=row, column=col)
            if cell.value == "R":
                cell.fill = openpyxl.styles.PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
    # Values with "(RA)" with text in dark red and italic
    for row in range(2, ws.max_row - 3):
        for col in range(2, ws.max_column - 4):
            cell = ws.cell(row=row, column=col)
            if "(RA)" in str(cell.value):
                cell.font = Font(color="8B0000", italic=True)
    # Last five columns in light grey
    for col in range(ws.max_column - 4, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    # Col Total in bold
    for row in range(1, ws.max_row - 3):
        ws.cell(row=row, column=ws.max_column - 4).font = bold_centered

    # Paint in orange the first shift of first agent
    multiple_6 = [j for j in J if j % 6 == 0]
    not_multiple_6 = [j for j in J if j % 6 != 0]
    nb_value_cols = 2 * len(multiple_6) + len(not_multiple_6)

    for i in I:
        ws.cell(row=2+i, column=(2 + (7 * (i - 1))) % nb_value_cols).fill = openpyxl.styles.PatternFill(
                                                                                                start_color="FFA500",
                                                                                                end_color="FFA500",
                                                                                                fill_type="solid"
                                                                                                )
    # Save the changes
    wb.save(dest_path)


def export_schedule(
        values, variable_names,
        I, J, K, part_time_I, nb_shifts,
        dest_path="output/nurses_schedule.xlsx"
        ):
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    to_excel(values, variable_names, I, J, K, part_time_I, nb_shifts, dest_path)
    openpyxl_formatting(I, J, dest_path, dest_path)
