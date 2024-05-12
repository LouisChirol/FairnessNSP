import cplex
from cplex.exceptions import CplexError
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
import pandas as pd

I = range(1, 12)  # Adjust according to your actual range  # noqa
J = range(1, 25)  # Adjust according to your actual range
K = range(1, 4)   # Adjust according to your actual range
part_time_I = range(1, 4)  # Indices of part-time agents in I


def populate_by_row(prob):
    prob.objective.set_sense(prob.objective.sense.minimize)

    # Add your variables
    # For a 3D variable x[i][j][k], you'll flatten this in a single list or array
    # E.g., x1,1,1, x1,1,2, ..., x2,12,3 (flattening by iterating over i, j, k)
    x_names = [f"x{i},{j},{k}" for i in I for j in J for k in K]
    prob.variables.add(names=x_names, types=["B"] * len(x_names))  # "I" denotes integer variables, "B" binary variables

    # Staffing constraints:
    variable_coefficients = [1] * len(I)  # Coefficient for each variable in the sum is 1
    for j in J:
        # First constraint (staff >= 3 in the morning shift):
        morning_variable_indices = [x_names.index(f"x{i},{j},1") for i in I]  # k=1 denotes morning shift
        prob.linear_constraints.add(
            lin_expr=[cplex.SparsePair(ind=morning_variable_indices, val=variable_coefficients)],
            senses=["G"],  # "G" for >= ; "L" for <= ; "E" for ==
            rhs=[3]  # Right hand side of the inequality
        )
        # Second constraint (staff >= 2 in the evening shift):
        evening_variable_indices = [x_names.index(f"x{i},{j},2") for i in I]  # k=2 denotes evening shift
        prob.linear_constraints.add(
            lin_expr=[cplex.SparsePair(ind=evening_variable_indices, val=variable_coefficients)],
            senses=["G"],
            rhs=[2]
        )
        # Third and fourth constraints (staff >= 1 in the day shift, exactly 0 for weekends):
        day_variable_indices = [x_names.index(f"x{i},{j},3") for i in I]  # k=3 denotes day shift
        if j % 6 != 0:  # TO VERIFY
            prob.linear_constraints.add(
                lin_expr=[cplex.SparsePair(ind=day_variable_indices, val=variable_coefficients)],
                senses=["G"],
                rhs=[1]
            )
        else:
            prob.linear_constraints.add(
                lin_expr=[cplex.SparsePair(ind=day_variable_indices, val=variable_coefficients)],
                senses=["E"],
                rhs=[0]
            )
        for i in range(1, 12):
            shift_indices = [x_names.index(f"x{i},{j},{k}") for k in K]
            # Fifth constaint (at most 1 shift per day for all agents):
            prob.linear_constraints.add(
                lin_expr=[cplex.SparsePair(ind=shift_indices, val=[1]*len(K))],
                senses=["L"],
                rhs=[1]
            )
            # Sixth constraint (>12h of rest between shifts):
            if j < max(J):
                shift_indices = [x_names.index(f"x{i},{j},2"), x_names.index(f"x{i},{j+1},1")]
                prob.linear_constraints.add(
                    lin_expr=[cplex.SparsePair(ind=shift_indices, val=[1, 1])],
                    senses=["L"],
                    rhs=[1]
                )
    for i in I:
        for j in range(1, max(J)//6):
            # Seventh constraint (week-end alternance):
            weekend_indices = ([x_names.index(f"x{i},{j*6},{k}") for k in K]  # saturday
                               + [x_names.index(f"x{i},{j*6+1},{k}") for k in K])  # sunday
            prob.linear_constraints.add(
                lin_expr=[cplex.SparsePair(ind=weekend_indices, val=[1]*len(weekend_indices))],
                senses=["E"],
                rhs=[1]
            )
        for j in range(0, max(J)//6):
            # Eighth constraint (2 days off per week):
            off_indices = [x_names.index(f"x{i},{jj},{k}") for jj in range(6*j+1, 6*(j+1)) for k in K]  # 6 days per week  # noqa
            prob.linear_constraints.add(
                lin_expr=[cplex.SparsePair(ind=off_indices, val=[1]*len(off_indices))],
                senses=["L"],
                rhs=[6*len(K)-2]
            )
        # Ninth constraint (9 rest days per month):
        month_off_indices = [x_names.index(f"x{i},{j},{k}") for j in J for k in K]
        prob.linear_constraints.add(
            lin_expr=[cplex.SparsePair(ind=month_off_indices, val=[-1]*len(month_off_indices))],
            senses=["L"],
            rhs=[len(J)*len(K)-9]
        )
    # Tenth constraint (off days for part-time agents):
    for i in part_time_I:
        part_time_indices = [x_names.index(f"x{i},{j},{k}") for j in J for k in K]
        prob.linear_constraints.add(
            lin_expr=[cplex.SparsePair(ind=part_time_indices, val=[1]*len(part_time_indices))],
            senses=["L"],
            rhs=[len(J)*len(K)-13]
        )
    # Eleventh constraint (at least 36h rest for off days):
    for i in I:
        for j in J[:-2]:
            off_indices = ([x_names.index(f"x{i},{j+1},{k}") for k in K]
                           + [x_names.index(f"x{i},{j},2"), x_names.index(f"x{i},{j+2},1")])
            off_variable_coefficients = [1] * len(K) + [-1, -1]
            prob.linear_constraints.add(
                lin_expr=[cplex.SparsePair(ind=off_indices, val=off_variable_coefficients)],
                senses=["G"],
                rhs=[-1]
            )

    # Define your objective function: minimize the total number of deviations to smooth the schedule
    # Auxiliary variables for deviations
    d_pos = [f"d_pos{i}" for i in I]  # Positive deviations for each agent
    d_neg = [f"d_neg{i}" for i in I]  # Negative deviations for each agent
    prob.variables.add(names=d_pos + d_neg,
                       lb=[0]*len(d_pos + d_neg),  # Lower bound of zero
                       types=["C"]*len(d_pos + d_neg))  # Continuous variables
    # Variable to capture the total sum of all x_ijk
    prob.variables.add(names=["total_var"],
                       lb=[0],  # Lower bound of zero, no upper bound needed
                       types=["I"])  # Integer variable

    # Objective function: Minimize sum of d_pos + d_neg
    prob.objective.set_sense(prob.objective.sense.minimize)
    obj_coeffs = [1] * (len(d_pos) + len(d_neg))
    prob.objective.set_linear(list(zip(d_pos + d_neg, obj_coeffs)))

    # Constraint to set total_var equal to the sum of all x_ijk
    total_sum_expr = cplex.SparsePair(ind=x_names + ["total_var"], val=[1]*len(x_names) + [-1])
    prob.linear_constraints.add(
        lin_expr=[total_sum_expr],
        senses=["E"],  # Equals sense
        rhs=[0]  # This ensures that total_var is the sum of all x_ijk
    )

    # Constraints to link deviations with x sums
    for i in I:
        sum_i_indices = [x_names.index(f"x{i},{j},{k}") for j in J for k in K]
        indices = sum_i_indices + [d_pos[i-1], d_neg[i-1], "total_var"]
        values = [1] * len(sum_i_indices) + [-1, 1, -1.0/len(I)]
        prob.linear_constraints.add(
            lin_expr=[cplex.SparsePair(ind=indices, val=values)],
            senses=["E"],
            rhs=[0]
        )


def to_excel(values):

    # Create the DataFrame structure
    data = []
    for i in I:  # For each nurse
        row = []
        for j in J:  # For each day
            for k in K:  # For each shift
                var_name = f"x{i},{j},{k}"
                index = variable_names.index(var_name)
                row.append(int(values[index]))  # Convert float to int for clarity
        data.append(row)

    # Convert list to DataFrame*
    shifts = {1: 'M', 2: 'S', 3: 'T'}
    days = {1: 'Monday', 2: 'Tuesday', 3: 'Wednesday', 4: 'Thursday', 5: 'Friday', 6: 'Week-end'}
    columns = [f"Week_{(j//3) // 6 + 1} {days[(j//3) % 6 + 1]} {shifts[k]}" for j in range(0, 72, 3) for k in K]
    df = pd.DataFrame(data, index=[f"Nurse {i}" for i in range(1, 12)], columns=columns)

    # Duplicate the week-end columns into two separate columns for Saturday and Sunday
    # First replace "Week-end" with "Saturday"
    df.columns = df.columns.str.replace('Week-end', 'Saturday')
    # Then duplicate the columns for Sunday, inserting them after Saturday
    weeks = max(J) // 6
    for w in range(weeks):
        for i, c in enumerate(df.columns):
            if 'Saturday' in c and f"Week_{w+1}" in c:
                df.insert(i+3, c.replace('Saturday', 'Sunday'), df[c])

    # Adding the total number of shifts per nurse
    df['Total Shifts'] = df.sum(axis=1)

    # Add a row for the number of nurses assigned per shift
    total_nurses = df.sum(axis=0)
    total_nurses.name = 'Total Nurses'
    df = pd.concat([df, total_nurses.to_frame().T])
    # Save as excel
    df.to_excel('nurse_schedule_v1.xlsx')


def openpyxl_formatting():

    # df = pd.read_excel("nurse_schedule.xlsx")
    wb = load_workbook("nurse_schedule_v1.xlsx")
    ws = wb.active

    # Assuming your data starts from column B (column A might be the index)
    # And that the 'Week X' formatting needs to apply from column B onward
    # Insert new rows for Week and Day headers
    ws.insert_rows(1)
    ws.insert_rows(1)

    # Apply header information
    header_row = 1  # Week headers
    day_header_row = 2  # Day headers

    # Set the initial week and day to manage merging
    current_week = None
    current_day = None
    week_start_col = 2
    day_start_col = 2

    for col in range(2, ws.max_column):
        # Extract week and day from the original cell
        original_header = ws.cell(row=3, column=col).value  # assuming original headers start from row 3 now
        if original_header:
            week, day, shift = original_header.split(" ")
            ws.cell(row=header_row, column=col).value = week
            ws.cell(row=day_header_row, column=col).value = day
            ws.cell(row=3, column=col).value = shift  # Shift goes back to row 3

            # Manage merging for week
            if week != current_week:
                if current_week is not None:
                    ws.merge_cells(start_row=header_row, start_column=week_start_col,
                                   end_row=header_row, end_column=col-1)
                current_week = week
                week_start_col = col

            # Manage merging for day
            if day != current_day:
                if current_day is not None:
                    ws.merge_cells(start_row=day_header_row, start_column=day_start_col,
                                   end_row=day_header_row, end_column=col-1)
                current_day = day
                day_start_col = col

    # Final merge for the last week and day
    if current_week:
        ws.merge_cells(start_row=header_row, start_column=week_start_col,
                       end_row=header_row, end_column=ws.max_column-1)
    if current_day:
        ws.merge_cells(start_row=day_header_row, start_column=day_start_col,
                       end_row=day_header_row, end_column=ws.max_column-1)

    # Set column widths and general styles as before
    for i, col in enumerate(ws.columns):
        # Skip first and last columns
        if i == 0 or i == ws.max_column-1:
            continue
        ws.column_dimensions[get_column_letter(col[0].column)].width = 5

    # Apply bold and centered style to day names
    bold_centered = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    for col in range(2, ws.max_column + 1):
        for row in [header_row, day_header_row]:
            cell = ws.cell(row=row, column=col)
            cell.font = bold_centered
            cell.alignment = center_alignment

    # Add a row with day-wise totals at the end
    total_row = ws.max_row
    new_row = ws.max_row + 1
    ws.insert_rows(new_row)
    ws.cell(row=new_row, column=1).value = "Total day-wise"
    ws.cell(row=new_row, column=1).font = bold_centered
    ws.cell(row=new_row, column=1).alignment = center_alignment
    for col in range(2, ws.max_column, 3):
        ws.cell(row=new_row, column=col).value = f"=SUM({get_column_letter(col)}{total_row}:{get_column_letter(col+2)}{total_row})"  # noqa
        ws.cell(row=new_row, column=col).font = bold_centered
        ws.cell(row=new_row, column=col).alignment = center_alignment
    # Group the total values for each day together
    for col in range(2, ws.max_column, 3):
        ws.merge_cells(start_row=new_row, start_column=col, end_row=new_row, end_column=col+2)

    # Define different border styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    # Apply thin borders to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Save the changes
    wb.save("nurse_schedule_openpyxl_v1.xlsx")


if __name__ == "__main__":
    try:
        my_prob = cplex.Cplex()
        populate_by_row(my_prob)

        # Print number of variables and constraints
        print("Variables: ", my_prob.variables.get_num())
        print("Constraints: ", my_prob.linear_constraints.get_num())
        print("Total: ", my_prob.variables.get_num() + my_prob.linear_constraints.get_num())

        print("Solving...")
        my_prob.solve()

    except CplexError as exc:
        print(exc)

    print("Solution status = ", my_prob.solution.get_status(), ":", end=' ')
    print(my_prob.solution.status[my_prob.solution.get_status()])
    print("Solution value  = ", my_prob.solution.get_objective_value())

    variable_names = [f"x{i},{j},{k}" for i in I for j in J for k in K]
    values = my_prob.solution.get_values(variable_names)

    to_excel(values)
    openpyxl_formatting()